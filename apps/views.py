import os, zipfile, qrcode, base64, re, tempfile
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.contrib import messages
from django.utils.text import slugify
from django.conf import settings
from django.views.generic import (
    TemplateView
)
from .forms import (
    PDFMergeForm, PDFSplitForm, QRCodeForm,
    BarcodeForm, PDFCompressForm
)
from io import BytesIO
from pypdf import PdfReader, PdfWriter
from PIL import Image

# Create your views here.
class PDFMerge(TemplateView):
    template_name = 'pages/form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Merge PDF'
        context['subtitle'] = 'Please upload at least two PDF files that you want to merge into a single document.'
        context['form'] = PDFMergeForm()
        context['buttons_action'] = [
                f'''
                <button type="submit" name="action" value="merge" class="btn-primary">
                    <span>Submit</span>
                    <i class="bi bi-arrow-right-circle"></i>
                </button>
                '''
            ]
        return context

    def post(self, request, *args, **kwargs):
        action = request.POST.get('action')
        referer = request.META.get('HTTP_REFERER', '/')

        if action == 'merge':
            files = request.FILES.getlist('pdf_files')
            
            if files:
                writer = PdfWriter()

                for f in files:
                    reader = PdfReader(f)
                    for page in reader.pages:
                        writer.add_page(page)

                merged_pdf = BytesIO()
                writer.write(merged_pdf)
                merged_pdf.seek(0)

                # Ambil nama file pertama sebagai dasar nama hasil
                first_filename = files[0].name
                base_name = os.path.splitext(first_filename)[0]
                merged_filename = f"{base_name}_merged.pdf"

                response = HttpResponse(merged_pdf, content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="{merged_filename}"'
                return response

        return redirect(referer)

class PDFSplit(TemplateView):
    template_name = 'pages/form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Split PDF'
        context['subtitle'] = 'Please upload the PDF file that you want to split into separate pages or sections.'
        context['form'] = PDFSplitForm()
        context['buttons_action'] = [
                f'''
                <button type="submit" name="action" value="split" class="btn-primary">
                    <span>Submit</span>
                    <i class="bi bi-arrow-right-circle"></i>
                </button>
                '''
            ]
        return context
    
    def post(self, request, *args, **kwargs):
        action = request.POST.get('action')
        referer = request.META.get('HTTP_REFERER', '/')

        if action == 'split':
            file = request.FILES.get('pdf_file')

            try:
                # Ambil nama file asli tanpa ekstensi
                original_filename = file.name
                base_filename = os.path.splitext(original_filename)[0]  # contoh: 'laporan_akhir'

                reader = PdfReader(file)
                zip_buffer = BytesIO()

                with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
                    for i, page in enumerate(reader.pages):
                        writer = PdfWriter()
                        writer.add_page(page)

                        pdf_bytes = BytesIO()
                        writer.write(pdf_bytes)
                        pdf_bytes.seek(0)

                        filename = f'page-{i+1}.pdf'
                        zip_file.writestr(filename, pdf_bytes.read())

                zip_buffer.seek(0)
                zip_filename = f"{base_filename}.zip"

                response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
                response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
                return response

            except Exception as e:
                messages.error(request,"PDF processing error:", e)
                return redirect(referer)

import fitz

def compress_pdf(input_file, dpi, quality):
    input_file.seek(0)
    doc = fitz.open(stream=input_file.read(), filetype="pdf")
    scale = dpi / 72  # Asumsikan DPI asli 72

    for page_index in range(len(doc)):
        page = doc[page_index]
        images = page.get_images(full=True)

        for img in images:
            xref = img[0]
            base_image = doc.extract_image(xref)
            image_bytes = base_image["image"]
            image_ext = base_image["ext"]

            if image_ext.lower() not in ["jpeg", "jpg", "png"]:
                continue

            try:
                # Load gambar asli dengan PIL
                image = Image.open(BytesIO(image_bytes)).convert("RGB")

                # Hitung ukuran baru berdasarkan dpi parameter
                new_width = int(image.width * scale)
                new_height = int(image.height * scale)

                # Resize gambar
                resized_image = image.resize((new_width, new_height), Image.LANCZOS)

                # Simpan ulang gambar sebagai JPEG dengan quality yang ditentukan
                img_buffer = BytesIO()
                resized_image.save(img_buffer, format="JPEG", quality=quality)
                new_image_bytes = img_buffer.getvalue()

                # Update gambar di PDF
                doc.update_stream(xref, new_image_bytes)

            except Exception as e:
                print(f"Gagal kompres gambar di halaman {page_index + 1}: {e}")

    output = BytesIO()
    doc.save(output, garbage=4, deflate=True, clean=True)
    doc.close()
    output.seek(0)
    return output

class PDFCompress(TemplateView):
    template_name = 'pages/form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Compress PDF'
        context['subtitle'] = 'Please upload the PDF file that you want to compress.'
        context['form'] = PDFCompressForm()
        context['buttons_action'] = [
                f'''
                <button type="submit" name="action" value="compress" class="btn-primary">
                    <span>Submit</span>
                    <i class="bi bi-arrow-right-circle"></i>
                </button>
                '''
            ]
        return context
    
    def post(self, request, *args, **kwargs):
        referer = request.META.get('HTTP_REFERER', '/')
        form = PDFCompressForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = form.cleaned_data['pdf_file']
            quality = form.cleaned_data['quality']

            dpi_mapping = {
                'high': 300,    # kualitas tinggi, cocok untuk cetak atau arsip
                'medium': 150,  # cukup tajam untuk layar, ukuran file kecil
                'low': 72,      # cukup untuk preview, teks kecil bisa blur
            }

            quality_value = {'high': 90, 'medium': 60, 'low': 30}[quality]
            dpi_value = dpi_mapping[quality]
            print(dpi_value)

            compressed_stream = compress_pdf(uploaded_file, dpi=dpi_value, quality=quality_value)
            response = HttpResponse(compressed_stream, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{uploaded_file.name.replace(".pdf", "")}_compressed.pdf"'
            return response
        return redirect(referer)

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

def pdf_to_excel(pdf_io):
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "PDF_Data"

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    current_row = 1
    max_col_lengths = {}

    with pdfplumber.open(pdf_io) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            # Ambil semua tabel dengan posisi bounding box
            pdf_tables = page.find_tables()

            # Simpan bounding box tabel (x0, top, x1, bottom)
            table_bboxes = [tbl.bbox for tbl in pdf_tables]

            # Ambil semua kata dengan posisi
            words = page.extract_words()

            # Filter kata yang TIDAK ada di area tabel manapun
            filtered_words = []
            for w in words:
                x0, top, x1, bottom = w['x0'], w['top'], w['x1'], w['bottom']
                inside_table = False
                for bbox in table_bboxes:
                    bx0, btop, bx1, bbottom = bbox
                    if (x0 >= bx0 and x1 <= bx1) and (top >= btop and bottom <= bbottom):
                        inside_table = True
                        break
                if not inside_table:
                    filtered_words.append(w)

            # Gabungkan kata filtered_words menjadi baris teks per posisi top
            lines = {}
            for w in filtered_words:
                top_key = round(w['top'], 1)
                lines.setdefault(top_key, []).append(w['text'])
            text_lines = [{"top": k, "text": " ".join(v)} for k, v in lines.items()]

            # Gabungkan semua objek halaman
            objects = []

            # Tambahkan teks luar tabel
            for tl in text_lines:
                objects.append({"type": "text", "top": tl["top"], "content": tl["text"]})

            # Tambahkan tabel dengan posisi
            for tbl in pdf_tables:
                bbox = tbl.bbox
                tables_content = tbl.extract()
                objects.append({"type": "table", "top": bbox[1], "content": tables_content})

            # Urutkan semua objek per posisi vertikal
            objects.sort(key=lambda x: x["top"])

            # Tulis isi objek urut
            for obj in objects:
                if obj["type"] == "text":
                    ws.cell(row=current_row, column=1, value=obj["content"])
                    current_row += 1
                elif obj["type"] == "table":
                    table = obj["content"]
                    for row in table:
                        col_idx = 1
                        c = 0
                        while c < len(row):
                            value = row[c] or ""
                            start_col = col_idx
                            # Merge horizontal sel kosong
                            end_col = start_col
                            for next_c in range(c+1, len(row)):
                                if row[next_c] is None or row[next_c] == "":
                                    end_col += 1
                                else:
                                    break
                            cell = ws.cell(row=current_row, column=start_col, value=value)
                            cell.border = thin_border
                            if end_col > start_col:
                                ws.merge_cells(start_row=current_row, start_column=start_col, end_row=current_row, end_column=end_col)
                            length = len(str(value))
                            for col_num in range(start_col, end_col+1):
                                if col_num not in max_col_lengths or length > max_col_lengths[col_num]:
                                    max_col_lengths[col_num] = length
                            col_idx = end_col + 1
                            c += (end_col - start_col + 1)
                        current_row += 1
                    current_row += 1  # spasi antar tabel

    # Atur lebar kolom
    for col_idx, max_len in max_col_lengths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(output)
    output.seek(0)
    return output

class PDFtoExcel(TemplateView):
    template_name = 'pages/form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Convert PDF to Excel'
        context['subtitle'] = 'Please upload the PDF file that you want to convert to an Excel spreadsheet.'
        context['form'] = PDFSplitForm()
        context['buttons_action'] = [
                f'''
                <button type="submit" name="action" value="convert" class="btn-primary">
                    <span>Submit</span>
                    <i class="bi bi-arrow-right-circle"></i>
                </button>
                '''
            ]
        return context
    
    def post(self, request, *args, **kwargs):
        action = request.POST.get('action')
        referer = request.META.get('HTTP_REFERER', '/')
        
        if action == 'convert':
            uploaded_file = request.FILES.get('pdf_file')
            # Gunakan BytesIO dari file upload langsung
            pdf_bytes = uploaded_file.read()
            pdf_io = BytesIO(pdf_bytes)
            excel_file = pdf_to_excel(pdf_io)
            response = HttpResponse(
                excel_file,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename={uploaded_file.name.replace(".pdf", "")}_{action}.xlsx'
            return response

        return redirect(referer)   

from pdf2docx import Converter

def pdf_to_word(pdf_io):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_pdf:
        temp_pdf.write(pdf_io.read())
        temp_pdf_path = temp_pdf.name

    with tempfile.NamedTemporaryFile(delete=False, suffix=".docx") as temp_docx:
        temp_docx_path = temp_docx.name

    cv = Converter(temp_pdf_path)
    cv.convert(temp_docx_path, start=0, end=None)
    cv.close()

    output = BytesIO()
    with open(temp_docx_path, "rb") as f:
        output.write(f.read())

    os.remove(temp_pdf_path)
    os.remove(temp_docx_path)

    output.seek(0)
    return output

class PDFtoWord(TemplateView):
    template_name = 'pages/form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Convert to Word'
        context['subtitle'] = 'Please upload the PDF file that you want to convert to a Word document.'
        context['form'] = PDFSplitForm()
        context['buttons_action'] = [
                f'''
                <button type="submit" name="action" value="convert" class="btn-primary">
                    <span>Submit</span>
                    <i class="bi bi-arrow-right-circle"></i>
                </button>
                '''
            ]
        return context
    
    def post(self, request, *args, **kwargs):
        action = request.POST.get('action')
        referer = request.META.get('HTTP_REFERER', '/')
        
        if action == 'convert':
            uploaded_file = request.FILES.get('pdf_file')
            if not uploaded_file:
                return HttpResponse("No file uploaded", status=400)

            # Baca PDF dari file upload ke BytesIO
            pdf_bytes = uploaded_file.read()
            pdf_io = BytesIO(pdf_bytes)

            # Konversi ke Word
            word_file = pdf_to_word(pdf_io)

            # Buat response HTTP
            response = HttpResponse(
                word_file,
                content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
            )
            response['Content-Disposition'] = f'attachment; filename="{uploaded_file.name.replace(".pdf", "")}_converted.docx"'
            return response
        return redirect(referer)

class QRCodeGenerator(TemplateView):
    template_name = 'pages/form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'QR Code Generator'
        context['subtitle'] = 'Input text, URL, or any information to create its QR Code.'
        context['form'] = QRCodeForm()
        context['buttons_action'] = [
                f'''
                <button type="submit" name="action" value="generate" class="btn-primary">
                    <span>Submit</span>
                    <i class="bi bi-arrow-right-circle"></i>
                </button>
                '''
            ]
        return context
    
    def post(self, request, *args, **kwargs):
        form = QRCodeForm(request.POST)
        context = self.get_context_data()

        if form.is_valid():
            data = form.cleaned_data['data']
            with_icon = form.cleaned_data.get('with_icon', False)  # Ambil nilai checkbox

            # Generate QR Code seperti biasa
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_H,
                box_size=10,
                border=4,
            )
            qr.add_data(data)
            qr.make(fit=True)

            img_qr = qr.make_image(fill_color="black", back_color="white").convert('RGB')

            if with_icon:
                # Jika checkbox dicentang, tambahkan logo
                logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo_01.jpg')
                logo = Image.open(logo_path)

                logo_size = int(img_qr.size[0] * 0.2)
                logo = logo.resize((logo_size, logo_size), Image.Resampling.LANCZOS)

                pos = ((img_qr.size[0] - logo_size) // 2, (img_qr.size[1] - logo_size) // 2)

                if logo.mode == 'RGBA':
                    img_qr.paste(logo, pos, mask=logo)
                else:
                    img_qr.paste(logo, pos)

            # Encode ke base64
            buffer = BytesIO()
            img_qr.save(buffer, format='PNG')
            img_str = base64.b64encode(buffer.getvalue()).decode('utf-8')
            img_data = f"data:image/png;base64,{img_str}"

            context['code_img'] = img_data
            context['text'] = data
            context['filename'] = slugify(data)

        context['form'] = form
        return render(request, self.template_name, context)
    
from barcode import Code128, EAN13
from barcode.writer import ImageWriter
from barcode import get_barcode_class

# Fungsi checksum Code 39 (Mod 43)
def checksum_mod43(data: str) -> str:
    charset = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ-. $/+%"
    total = sum(charset.index(c) for c in data if c in charset)
    return charset[total % 43]

class BarcodeGenerator(TemplateView):
    template_name = 'pages/form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Barcode Generator'
        context['subtitle'] = 'Please enter the text or data you want to convert into a barcode.'
        context['form'] = BarcodeForm()
        context['buttons_action'] = [
                f'''
                <button type="submit" name="action" value="generate" class="btn-primary">
                    <span>Submit</span>
                    <i class="bi bi-arrow-right-circle"></i>
                </button>
                '''
            ]
        return context
    
    def post(self, request, *args, **kwargs):
        form = BarcodeForm(request.POST)
        context = self.get_context_data()

        if form.is_valid():
            data = form.cleaned_data['data']
            barcode_type = form.cleaned_data['barcode_type']

            if barcode_type in ['ean13', 'isbn13', 'upc'] and (not data.isdigit() or len(data) != 12):
                messages.error(request, f'{barcode_type.upper()} must be exactly 12 numeric digits.')
                form.add_error('data', f'{barcode_type.upper()} must be exactly 12 numeric digits.')
            elif barcode_type == 'ean8' and (not data.isdigit() or len(data) != 7):
                messages.error(request, 'EAN-8 must be exactly 7 numeric digits.')
                form.add_error('data', 'EAN-8 must be exactly 7 numeric digits.')
            elif barcode_type == 'isbn10' and (not data.isdigit() or len(data) != 9):
                messages.error(request, 'ISBN-10 must be exactly 9 numeric digits.')
                form.add_error('data', 'ISBN-10 must be exactly 9 numeric digits.')
            elif barcode_type == 'issn' and (not data.isdigit() or len(data) != 7):
                messages.error(request, 'ISSN must be exactly 7 numeric digits.')
                form.add_error('data', 'ISSN must be exactly 7 numeric digits.')
            elif barcode_type in ['code39', 'code39_mod43']:
                allowed_chars = r'[0-9A-Z\-\.\ \$\/\+\%]*'
                if not re.fullmatch(allowed_chars, data):
                    messages.error(request, 'Code 39 only allows uppercase letters, digits, and - . $ / + % (space).')
                    form.add_error('data', 'Code 39 only allows uppercase letters, digits, and - . $ / + % (space).')
                else:
                    # Tambahkan checksum jika kode mod43
                    if barcode_type == 'code39_mod43':
                        data += checksum_mod43(data)

                    try:
                        print(data)
                        barcode_class = get_barcode_class('code39')  # Tetap pakai 'code39'
                        barcode = barcode_class(data, writer=ImageWriter(), add_checksum=False)
                        buffer = BytesIO()
                        barcode.write(buffer)
                        img_str = base64.b64encode(buffer.getvalue()).decode()
                        context['code_img'] = f"data:image/png;base64,{img_str}"
                    except Exception as e:
                        form.add_error('data', f'Barcode generation failed: {str(e)}')

            else:
                # General barcode handler
                try:
                    barcode_class = get_barcode_class(barcode_type)
                    barcode = barcode_class(data, writer=ImageWriter())
                    buffer = BytesIO()
                    barcode.write(buffer)
                    img_str = base64.b64encode(buffer.getvalue()).decode()
                    context['code_img'] = f"data:image/png;base64,{img_str}"
                except Exception as e:
                    form.add_error('data', f'Failed to generate barcode: {str(e)}')

        context['form'] = form
        context['text'] = data
        context['filename'] = slugify(data)
        return render(request, self.template_name, context)
